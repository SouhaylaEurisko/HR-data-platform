"""
Utility script to debug how the XLSX import logic sees sheets in the Excel file.

It uses the same table‑detection helpers as the main import pipeline so we can:
- list how many tables are detected in the sheet
- see the header range (columns) for each table
- inspect a few sample data rows per table as key/value dictionaries

This is meant for local debugging and is not wired into any API.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

# Re‑use the helpers from the import service so behaviour matches the API
from ..services.import_service import _detect_tables_in_sheet  # type: ignore[attr-defined]


DATA_FILE = Path(
    "/storage/souhayla/project1/data/Pre- Screening and Follow-up (1).xlsx"
)
SHEETS_TO_TEST = ["ReactNative", "Java"]


def _row_to_dict(
    row_values: tuple,
    headers: List[str],
    min_col: int,
    max_col: int,
) -> Dict[str, Any]:
    """
    Convert a raw XLSX row into a {header: value} dict using the same
    column slicing logic as the import pipeline.
    """
    # Slice to table column range, guarding against short rows
    actual_max_col = min(max_col + 1, len(row_values))
    sliced = row_values[min_col:actual_max_col]

    row_dict: Dict[str, Any] = {}
    num_cols = min(len(headers), len(sliced))

    for i in range(num_cols):
        header = headers[i]
        value = sliced[i]
        if header and str(header).strip():
            key = str(header).strip()
            counter = 1
            while key in row_dict:
                key = f"{header}_{counter}"
                counter += 1
            row_dict[key] = value

    return row_dict


def analyse_sheet(sheet_name: str, wb) -> None:
    """Analyze a single sheet."""
    if sheet_name not in wb.sheetnames:
        print(f"⚠️  Sheet {sheet_name!r} not found in workbook.")
        print(f"   Available sheets: {', '.join(wb.sheetnames)}")
        return

    sheet = wb[sheet_name]

    print(f"\n{'='*80}")
    print(f"Analysing file: {DATA_FILE}")
    print(f"Sheet: {sheet_name}")
    print(f"Max row: {sheet.max_row}, max column: {sheet.max_column}")
    print("-" * 80)

    tables = _detect_tables_in_sheet(sheet)
    print(f"Detected tables: {len(tables)}")
    print()

    for idx, t in enumerate(tables):
        start_row = t["start_row"]
        end_row = t["end_row"] or sheet.max_row
        headers = t["headers"]
        min_col = t.get("min_col", 0)
        max_col = t.get("max_col", sheet.max_column - 1)

        print(f"Table #{idx}")
        print(f"  Rows: {start_row} -> {end_row}")
        print(f"  Cols: {min_col + 1} -> {max_col + 1}")
        print(f"  Header cells: {headers}")

        # Collect up to 5 non‑empty data rows for inspection
        samples: List[Dict[str, Any]] = []
        for excel_row_idx, row in enumerate(
            sheet.iter_rows(
                min_row=start_row + 1,
                max_row=end_row,
                values_only=True,
            ),
            start=start_row + 1,
        ):
            row_dict = _row_to_dict(row, headers, min_col, max_col)
            if any(v not in (None, "") for v in row_dict.values()):
                samples.append({"row_index": excel_row_idx, "data": row_dict})
            if len(samples) >= 5:
                break

        if not samples:
            print("  Sample rows: (none, all empty within this range)")
        else:
            print("  Sample rows:")
            for sample in samples:
                print(f"    - Excel row {sample['row_index']}: {sample['data']}")
        print("-" * 80)


def analyse_sheets() -> None:
    """Analyze multiple sheets."""
    if not DATA_FILE.exists():
        raise FileNotFoundError(f"Data file not found: {DATA_FILE}")

    wb = load_workbook(filename=str(DATA_FILE), data_only=True)
    
    for sheet_name in SHEETS_TO_TEST:
        analyse_sheet(sheet_name, wb)


if __name__ == "__main__":
    analyse_sheets()

