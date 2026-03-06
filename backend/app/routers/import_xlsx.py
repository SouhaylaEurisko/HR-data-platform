from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..db import get_db
from ..models import Candidate
from ..schemas import CandidateCreate

router = APIRouter(prefix="/api/import", tags=["import"])


class FieldMappingConfig:
    """
    Configuration for mapping Excel column headers to Candidate fields.
    Supports common fields and sheet-specific overrides.
    """
    
    # Common field mappings - used for all sheets
    COMMON_FIELDS: Dict[str, List[str]] = {
        "full_name": ["candidate name", "full name", "applicant name", "fullname", "name", "candidate"],
        "email": ["email", "e-mail", "email address", "e mail"],
        "nationality": ["nationality", "country of origin", "country of residence", "country"],
        "date_of_birth": ["date of birth", "birth date", "dob", "birthday", "date of birth (dd/mm/yyyy)"],
        "position": ["position", "job title", "role", "position applying for", "applied position", "job"],
        "expected_salary": [
            "expected salary",
            "salary expectations",
            "salary expectation",
            "desired salary",
            "salary",
            "expected salary (usd)",
        ],
        "years_experience": [
            "years of experience",
            "total experience (years)",
            "total years of experience",
            "in total, how many years of experience do you have?",
            "experience",
            "yrs of experience",
            "YOE",
            "total experience",
        ],
        "current_address": [
            "current address",
            "current location",
            "current residence",
            "location",
            "address",
            "city",
        ],
        "notice_period": [
            "notice period",
            "notice period (days)",
            "notice period (weeks)",
            "notice period (months)",
            "availability",
            "when can you start?",
        ],
    }
    
    # Sheet-specific overrides - maps sheet names to field overrides
    SHEET_OVERRIDES: Dict[str, Dict[str, List[str]]] = {
        # Example: "Pre-Screening": {
        #     "full_name": ["candidate full name", "applicant"],
        #     "position": ["role applied for"],
        # },
    }
    
    @classmethod
    def get_field_mappings(cls, sheet_name: Optional[str] = None) -> Dict[str, List[str]]:
        """
        Get field mappings for a specific sheet.
        Merges common fields with sheet-specific overrides.
        """
        mappings = cls.COMMON_FIELDS.copy()
        
        if sheet_name and sheet_name in cls.SHEET_OVERRIDES:
            # Merge sheet-specific overrides
            for field, headers in cls.SHEET_OVERRIDES[sheet_name].items():
                if field in mappings:
                    # Combine with common mappings, sheet-specific first
                    mappings[field] = headers + [h for h in mappings[field] if h not in headers]
                else:
                    mappings[field] = headers
        
        return mappings


async def _load_workbook_from_file(file: UploadFile):
    """
    Common function to validate and load workbook from uploaded file.
    """
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx files are supported for this endpoint.",
        )

    try:
        contents = await file.read()
        from io import BytesIO
        workbook = load_workbook(filename=BytesIO(contents), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to read XLSX file: {exc}",
        ) from exc

    if not workbook.sheetnames:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The provided XLSX file has no sheets.",
        )

    return workbook


def _to_json_safe(value: Any) -> Any:
    """
    Convert cell values to JSON-serializable primitives.
    - datetime/date → ISO string
    - other non-serializable types → str(value)
    """
    from datetime import date, datetime

    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    # Fallback: string representation
    return str(value)


def _to_float_or_none(value: Any) -> float | None:
    """
    Try to convert a cell value to float.
    Handles numeric types directly and simple strings (optionally with text).
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None

    import re

    # Handle numeric ranges like "1800-2000" or "1,800 - 2,000"
    range_match = re.match(r"^\s*([\d\.,]+)\s*[-–]\s*([\d\.,]+)\s*$", text)
    if range_match:
        first, second = range_match.groups()

        def _to_single_float(part: str) -> float:
            cleaned_part = re.sub(r"[^0-9\.,]", "", part)
            if not cleaned_part:
                raise ValueError("Empty numeric part in range")
            cleaned_part = cleaned_part.replace(",", ".")
            return float(cleaned_part)

        try:
            lower = _to_single_float(first)
            upper = _to_single_float(second)
            # Use the lower bound for numeric filtering / sorting
            return min(lower, upper)
        except ValueError:
            # Fall back to generic parsing below
            pass

    # Generic numeric parsing
    cleaned = re.sub(r"[^0-9\.,]", "", text)
    if not cleaned:
        return None
    cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _to_date_or_none(value: Any):
    """
    Convert an XLSX cell value to a Python date (for the DB).
    Accepts:
    - date/datetime objects (from openpyxl)
    - ISO-like strings (YYYY-MM-DD)
    """
    from datetime import date, datetime

    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()

    text = str(value).strip()
    if not text:
        return None

    # Try ISO format first, including variants with time component, e.g. "1997-12-29T00:00:00"
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        pass

    # Try a couple common explicit formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _is_empty_row(row: tuple) -> bool:
    """Check if a row is completely empty."""
    return all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row)


def _is_header_row(row: tuple) -> bool:
    """
    Determine if a row looks like a header row.

    Updated, tighter heuristic:
    - Requires at least 3 non-empty cells.
    - Looks for cells whose text matches known header patterns from FieldMappingConfig.
    - Ignores cells that contain digits (to avoid matching data like \"Current Salary 1500$\").
    - Treats a row as a header only if it contains at least 2 distinct header-like cells.
    """
    # Build a flat list of known header patterns from the field mapping config
    header_patterns: List[str] = []
    for patterns in FieldMappingConfig.COMMON_FIELDS.values():
        header_patterns.extend(p.lower() for p in patterns)

    non_empty = [cell for cell in row if cell is not None and str(cell).strip()]
    if len(non_empty) < 3:
        return False

    header_like_count = 0
    seen_patterns: set[str] = set()

    for cell in non_empty:
        if not isinstance(cell, str):
            continue
        text = cell.strip().lower()
        if not text:
            continue

        # If the cell contains digits, it's likely data (dates, numeric values, etc.)
        # Skip these for header detection.
        if any(ch.isdigit() for ch in text):
            continue

        for pattern in header_patterns:
            if pattern in text:
                if pattern not in seen_patterns:
                    seen_patterns.add(pattern)
                    header_like_count += 1
                break

    # Require at least 2 distinct header-like patterns to consider this a header row.
    return header_like_count >= 2


def _detect_tables_in_sheet(sheet) -> List[dict]:
    """
    Detect multiple tables within a single sheet.
    
    Returns list of table info dictionaries:
    - start_row: First row of table (header row index, 1-based)
    - end_row: Last row of table (or None if until end)
    - headers: List of header names
    """
    tables = []
    current_table_start = None
    current_headers = None
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # Check if this row looks like headers
        if _is_header_row(row):
            # If we had a previous table, close it
            if current_table_start:
                tables.append({
                    "start_row": current_table_start,
                    "end_row": row_idx - 1,
                    "headers": current_headers,
                })
            
            # Start new table
            current_table_start = row_idx
            current_headers = [str(cell).strip() if cell is not None else "" for cell in row]
        
        # Check for empty row (potential table separator)
        elif _is_empty_row(row) and current_table_start:
            # Only close table if we've seen at least one data row after headers
            if row_idx > current_table_start + 1:
                tables.append({
                    "start_row": current_table_start,
                    "end_row": row_idx - 1,
                    "headers": current_headers,
                })
                current_table_start = None
                current_headers = None
    
    # Close last table if exists
    if current_table_start:
        tables.append({
            "start_row": current_table_start,
            "end_row": None,  # Until end of sheet
            "headers": current_headers,
        })
    
    return tables


def _find_field_value(normalized_headers: Dict[str, Any], possible_headers: List[str]) -> Optional[str]:
    """
    Find a field value by matching against possible header names.
    Prioritizes exact matches, then substring matches.
    
    Args:
        normalized_headers: Dict of normalized header -> value
        possible_headers: List of possible header names to match
    
    Returns:
        First matching value as string, or None
    """
    # First pass: Try exact matches (most precise)
    for possible in possible_headers:
        possible_lower = possible.lower()
        if possible_lower in normalized_headers:
            value = normalized_headers[possible_lower]
            if value not in (None, ""):
                return str(value).strip()
    
    # Second pass: Try substring matches (header contains the pattern)
    # This handles cases like "candidate name" matching "name"
    for possible in possible_headers:
        possible_lower = possible.lower()
        for header, value in normalized_headers.items():
            if value in (None, ""):
                continue
            # Only match if the pattern is in the header (not the other way around)
            # This prevents "date" from matching "name" or vice versa
            if possible_lower in header:
                return str(value).strip()
    
    return None


def _map_row_to_candidate_create(
    row_data: dict,
    source_file: str,
    source_sheet: str,
    source_table_index: int,
    row_index: int,
    mapping_config: Optional[FieldMappingConfig] = None,
) -> CandidateCreate:
    """
    Map XLSX row data to CandidateCreate using dynamic field mapping.
    
    Args:
        row_data: Raw row data as dict
        source_file: Source filename
        source_sheet: Source sheet name
        source_table_index: Table index within sheet
        row_index: Row index (1-based)
        mapping_config: Optional field mapping configuration (uses default if None)
    
    Returns:
        CandidateCreate instance with mapped fields
    """
    if mapping_config is None:
        mapping_config = FieldMappingConfig()
    
    # Normalize headers for easier matching
    normalized_headers = {
        k.strip().lower(): _to_json_safe(v)
        for k, v in row_data.items()
        if k is not None
    }
    
    # Get field mappings for this sheet
    field_mappings = mapping_config.get_field_mappings(source_sheet)
    
    # Map common fields
    mapped_data = {}
    for field_name, possible_headers in field_mappings.items():
        value = _find_field_value(normalized_headers, possible_headers)
        if value:
            mapped_data[field_name] = value
    
    # Store all original data in raw_data (all fields are preserved)
    raw_data = row_data.copy()
    
    # Apply type conversions
    date_of_birth = _to_date_or_none(mapped_data.get("date_of_birth"))

    # Normalize email: if it doesn't look like a valid email, drop it (set to None)
    raw_email = mapped_data.get("email")
    email_value: Optional[str] = None
    if raw_email:
        email_str = str(raw_email).strip()
        if "@" in email_str and "." in email_str:
            email_value = email_str

    raw_expected_salary = mapped_data.get("expected_salary")
    expected_salary = _to_float_or_none(raw_expected_salary)
    # Preserve the original text for display (e.g., ranges like "1800-2000")
    expected_salary_text = raw_expected_salary if raw_expected_salary is not None else None

    years_experience = _to_float_or_none(mapped_data.get("years_experience"))
    
    return CandidateCreate(
        source_file=source_file,
        source_sheet=source_sheet,
        source_table_index=source_table_index,
        row_index=row_index,
        full_name=mapped_data.get("full_name"),
        email=email_value,
        nationality=mapped_data.get("nationality"),
        date_of_birth=date_of_birth,
        position=mapped_data.get("position"),
        expected_salary=expected_salary,
        expected_salary_text=expected_salary_text,
        current_address=mapped_data.get("current_address"),
        years_experience=years_experience,
        notice_period=mapped_data.get("notice_period"),
        raw_data=raw_data,
    )


def _process_table(
    sheet,
    table_info: dict,
    source_file: str,
    source_sheet: str,
    source_table_index: int,
    db: Session,
) -> dict:
    """
    Process a single table within a sheet.
    
    Args:
        sheet: The worksheet object
        table_info: Dict with 'start_row', 'end_row', 'headers'
        source_file: Source filename
        source_sheet: Source sheet name
        source_table_index: Index of this table (0-based)
        db: Database session
    """
    start_row = table_info["start_row"]
    end_row = table_info["end_row"]
    headers = table_info["headers"]
    
    created_count = 0
    skipped_empty = 0
    skipped_duplicates = 0
    row_errors: list[dict] = []
    
    # Determine row range
    if end_row is None:
        # Process until end of sheet
        rows_to_process = sheet.iter_rows(min_row=start_row + 1, values_only=True)
    else:
        rows_to_process = sheet.iter_rows(min_row=start_row + 1, max_row=end_row, values_only=True)
    
    for row_idx, row in enumerate(rows_to_process, start=start_row + 1):
        # Build row dictionary - headers and values must align by column index
        # This ensures each value is paired with its corresponding header by position
        row_dict = {}
        num_cols = min(len(headers), len(row))
        
        for i in range(num_cols):
            header = headers[i]
            value = row[i]
            
            # Only include columns with non-empty headers
            # This preserves column alignment - empty headers are skipped but indices stay aligned
            if header and header.strip():
                # Handle duplicate headers by appending column index to make keys unique
                header_key = header
                counter = 1
                while header_key in row_dict:
                    header_key = f"{header}_{counter}"
                    counter += 1
                row_dict[header_key] = _to_json_safe(value)
        
        # Skip entirely empty rows
        if not any(value not in (None, "") for value in row_dict.values()):
            skipped_empty += 1
            continue
        
        try:
            candidate_create = _map_row_to_candidate_create(
                row_data=row_dict,
                source_file=source_file,
                source_sheet=source_sheet,
                source_table_index=source_table_index,
                row_index=row_idx,
                mapping_config=FieldMappingConfig(),
            )
            
            # Check for duplicate using email as the unique identifier
            if candidate_create.email:
                existing = db.query(Candidate).filter(
                    Candidate.email == candidate_create.email
                ).first()
                if existing:
                    skipped_duplicates += 1
                    continue
            
            # Create candidate from schema using dict unpacking
            candidate = Candidate(**candidate_create.model_dump())
            db.add(candidate)
            created_count += 1
        except Exception as exc:  # noqa: BLE001
            row_errors.append({
                "row_index": row_idx,
                "error": str(exc),
            })
    
    return {
        "created": created_count,
        "skipped_empty_rows": skipped_empty,
        "skipped_duplicates": skipped_duplicates,
        "row_errors": row_errors,
    }


def _process_sheet(
    sheet,
    source_file: str,
    source_sheet: str,
    db: Session,
) -> dict:
    """
    Process a single sheet, detecting and handling multiple tables.
    """
    # Detect tables in the sheet
    tables = _detect_tables_in_sheet(sheet)
    
    # If no tables detected, treat entire sheet as one table (backward compatible)
    if not tables:
        # Fallback: treat first row as headers
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            headers = next(rows_iter)
            header_list = [str(h).strip() if h is not None else "" for h in headers]
            tables = [{
                "start_row": 1,
                "end_row": None,
                "headers": header_list,
            }]
        except StopIteration:
            return {
                "sheet_name": source_sheet,
                "tables_found": 0,
                "created": 0,
                "skipped_empty_rows": 0,
                "skipped_duplicates": 0,
                "row_errors": [],
                "table_results": [],
            }
    
    # Process each table
    sheet_created = 0
    sheet_skipped_empty = 0
    sheet_skipped_duplicates = 0
    sheet_row_errors = []
    table_results = []
    
    for table_idx, table_info in enumerate(tables):
        table_result = _process_table(
            sheet=sheet,
            table_info=table_info,
            source_file=source_file,
            source_sheet=source_sheet,
            source_table_index=table_idx,
            db=db,
        )
        
        sheet_created += table_result["created"]
        sheet_skipped_empty += table_result["skipped_empty_rows"]
        sheet_skipped_duplicates += table_result["skipped_duplicates"]
        sheet_row_errors.extend(table_result["row_errors"])
        
        table_results.append({
            "table_index": table_idx,
            "start_row": table_info["start_row"],
            "end_row": table_info["end_row"],
            "created": table_result["created"],
            "skipped_empty_rows": table_result["skipped_empty_rows"],
            "skipped_duplicates": table_result["skipped_duplicates"],
            "row_errors": table_result["row_errors"],
        })
    
    return {
        "sheet_name": source_sheet,
        "tables_found": len(tables),
        "created": sheet_created,
        "skipped_empty_rows": sheet_skipped_empty,
        "skipped_duplicates": sheet_skipped_duplicates,
        "row_errors": sheet_row_errors,
        "table_results": table_results,
    }


def _determine_sheets_to_process(
    workbook, sheet_names: Optional[List[str]], import_all_sheets: bool
) -> List[str]:
    """
    Determine which sheets to process based on parameters.
    """
    all_sheet_names = workbook.sheetnames

    if import_all_sheets:
        # Explicit request: import all sheets
        return all_sheet_names

    if sheet_names:
        # Validate that requested sheets exist
        valid_sheets = [s for s in sheet_names if s in all_sheet_names]
        if not valid_sheets:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"None of the requested sheets found. Available sheets: {', '.join(all_sheet_names)}",
            )
        return valid_sheets

    # Default behavior: import all sheets when no specific selection is provided
    return all_sheet_names


@router.post("/xlsx/preview", summary="Preview XLSX file structure")
async def preview_xlsx(
    file: UploadFile = File(...),
) -> dict:
    """
    Preview an XLSX file structure without importing.
    Returns list of sheets and their basic information.
    """
    workbook = await _load_workbook_from_file(file)

    sheets_info = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        max_row = sheet.max_row
        non_empty_rows = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):
                non_empty_rows += 1
        
        sheets_info.append({
            "name": sheet_name,
            "max_row": max_row,
            "data_rows": non_empty_rows,
        })

    return {
        "file_name": file.filename or "unknown.xlsx",
        "sheets": sheets_info,
        "total_sheets": len(sheets_info),
    }


@router.post("/xlsx", summary="Upload and import an XLSX file")
async def import_xlsx(
    file: UploadFile = File(...),
    sheet_names: Optional[List[str]] = Query(None, description="Specific sheet names to import"),
    import_all_sheets: bool = Query(False, description="Import all sheets in the workbook"),
    db: Session = Depends(get_db),
) -> dict:
    """
    Accepts an XLSX file and imports candidate data from one or more sheets.

    - By default, imports **all sheets** when no specific selection is provided.
    - Use `import_all_sheets=true` to explicitly import all sheets.
    - Use `sheet_names` query parameter to import specific sheets (e.g., ?sheet_names=Sheet1&sheet_names=Sheet2).

    For each sheet:
    - Automatically detects multiple tables within sheets.
    - Assumes the first row of each table contains headers.
    - Each subsequent row becomes one candidate.
    - Uses a heuristic mapping from column names to Candidate fields.
    """
    workbook = await _load_workbook_from_file(file)

    # Determine which sheets to process
    sheets_to_process = _determine_sheets_to_process(
        workbook, sheet_names, import_all_sheets
    )

    # Process each sheet
    sheet_results = []
    total_created = 0
    total_skipped_empty = 0
    total_skipped_duplicates = 0
    all_row_errors = []

    for sheet_name in sheets_to_process:
        sheet = workbook[sheet_name]
        sheet_result = _process_sheet(
            sheet=sheet,
            source_file=file.filename or "unknown.xlsx",
            source_sheet=sheet_name,
            db=db,
        )
        sheet_results.append(sheet_result)
        total_created += sheet_result["created"]
        total_skipped_empty += sheet_result["skipped_empty_rows"]
        total_skipped_duplicates += sheet_result["skipped_duplicates"]
        # Add sheet name to each error for better tracking
        all_row_errors.extend(
            {**error, "sheet": sheet_name} for error in sheet_result["row_errors"]
        )

    db.commit()

    # Build response
    response = {
        "file_name": file.filename or "unknown.xlsx",
        "sheets_processed": sheets_to_process,
        "total_created": total_created,
        "total_skipped_empty_rows": total_skipped_empty,
        "total_skipped_duplicates": total_skipped_duplicates,
        "sheet_results": sheet_results,
    }

    # Include row_errors only if there are any (for backward compatibility)
    if all_row_errors:
        response["row_errors"] = all_row_errors

    return response

