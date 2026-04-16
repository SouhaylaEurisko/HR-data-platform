"""
Import service — smart XLSX import pipeline with LLM column normalization,
programmatic validation, lookup resolution, and single-INSERT writes.
"""

import os
import re
import tempfile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Dict, List, Optional, Set, Protocol

from fastapi import UploadFile
from openpyxl import load_workbook

from ..data.import_column_sets import (
    APPLICATION_IMPORT_KEYS,
    BOOLEAN_COLUMNS,
    DATE_COLUMNS,
    DECIMAL_COLUMNS,
    INT_COLUMNS,
    LOOKUP_COLUMN_MAP,
    STRING_COLUMNS,
)
from ..models.enums import RelocationOpenness, TransportationAvailability
from ..repository.import_repository import ImportRepositoryProtocol
from .column_normalizer_service import (
    ColumnNormalizerServiceProtocol,
    ColumnMapping,
    COLUMN_LABELS,
)
from .import_header_detection_service import detect_all_header_rows
from .lookup_service import LookupServiceProtocol
from .custom_field_type_service import (
    CustomFieldTypeServiceProtocol,
)


# ──────────────────────────────────────────────
# Temp file storage for two-phase import
# ──────────────────────────────────────────────

_UPLOAD_DIR = os.path.join(tempfile.gettempdir(), "hr_import_uploads")
os.makedirs(_UPLOAD_DIR, exist_ok=True)


def _save_temp_file(session_id: int, contents: bytes) -> str:
    path = os.path.join(_UPLOAD_DIR, f"session_{session_id}.xlsx")
    with open(path, "wb") as f:
        f.write(contents)
    return path


def _load_temp_file(session_id: int):
    path = os.path.join(_UPLOAD_DIR, f"session_{session_id}.xlsx")
    if not os.path.exists(path):
        raise FileNotFoundError(f"Temp file for session {session_id} not found.")
    return load_workbook(filename=path, data_only=True)


def _cleanup_temp_file(session_id: int):
    path = os.path.join(_UPLOAD_DIR, f"session_{session_id}.xlsx")
    if os.path.exists(path):
        os.remove(path)


# ──────────────────────────────────────────────
# Value conversion helpers
# ──────────────────────────────────────────────

def _to_json_safe(value: Any) -> Any:
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return str(value)


def _normalize_numeric_string_to_decimal_str(cleaned: str) -> Optional[str]:
    """
    Turn a cleaned string (digits, optional . and , and leading -) into a form
    Decimal() accepts.
    """
    s = cleaned.strip()
    if not s:
        return None
    neg = False
    if s.startswith("-"):
        neg = True
        s = s[1:]
    if not s or not re.fullmatch(r"[\d\.,]+", s):
        return None

    def with_sign(num: str) -> str:
        return ("-" if neg else "") + num

    if "," in s and "." in s:
        # Last separator is the decimal separator
        if s.rfind(",") > s.rfind("."):
            # e.g. 1.234,56
            num = s.replace(".", "").replace(",", ".")
        else:
            # e.g. 1,234.56
            num = s.replace(",", "")
        return with_sign(num)

    if "," in s:
        parts = s.split(",")
        # Multiple groups, last segment short → decimal comma (1234,5 / 1234,56)
        if len(parts) > 1 and len(parts[-1]) <= 2:
            num = "".join(parts[:-1]) + "." + parts[-1]
        else:
            # Thousands only: 5,000
            num = "".join(parts)
        return with_sign(num)

    if "." in s:
        parts = s.split(".")
        # Dot as thousands: every segment after the first is exactly 3 digits
        if len(parts) >= 2 and all(len(p) == 3 for p in parts[1:]):
            return with_sign("".join(parts))
        return with_sign(s)

    return with_sign(s)


def _to_decimal_or_none(value: Any) -> Optional[Decimal]:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        return None
    cleaned = re.sub(r"[^0-9\.,\-]", "", text)
    if not cleaned:
        return None
    # Handle ranges: take the lower bound
    range_match = re.match(r"^([\d\.,]+)\s*[-–]\s*([\d\.,]+)$", cleaned)
    if range_match:
        cleaned = range_match.group(1)
    normalized = _normalize_numeric_string_to_decimal_str(cleaned)
    if not normalized:
        return None
    try:
        return Decimal(normalized)
    except InvalidOperation:
        return None


def _to_date_or_none(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _to_bool_or_none(value: Any) -> Optional[bool]:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in ("true", "yes", "y", "employed"):
        return True
    if text in ("false", "no", "n", "unemployed", "not employed"):
        return False
    return None


def _to_relocation_openness_or_none(value: Any) -> Optional[RelocationOpenness]:
    """Map Excel/form values to RelocationOpenness (yes | no | for_missions_only)."""
    if value is None:
        return None
    if isinstance(value, RelocationOpenness):
        return value
    text = str(value).strip().lower()
    if not text:
        return None
    # Third option (Google Forms / HR wording)
    if "mission" in text:
        return RelocationOpenness.for_missions_only
    if text in ("true", "yes","y"):
        return RelocationOpenness.yes
    if text in ("false", "no", "n"):
        return RelocationOpenness.no
    # Try enum value / name
    for member in RelocationOpenness:
        if text == member.value or text == member.name.lower():
            return member
    return None


def _to_transportation_availability_or_none(
    value: Any,
) -> Optional[TransportationAvailability]:
    """Map Excel/form values to TransportationAvailability (yes | no | only_open_for_remote_opportunities)."""
    if value is None:
        return None
    if isinstance(value, TransportationAvailability):
        return value
    text = str(value).strip().lower()
    if not text:
        return None
    for member in TransportationAvailability:
        if text == member.value or text == member.name.lower():
            return member
    if text in ("true", "yes", "has transportation", "has car", "own vehicle"):
        return TransportationAvailability.yes
    if text in ("false", "no","no transportation"):
        return TransportationAvailability.no
    if "only" in text and "remote" in text:
        return TransportationAvailability.only_open_for_remote_opportunities
    if "remote only" in text or "remote opportunities" in text:
        return TransportationAvailability.only_open_for_remote_opportunities
    return None


def _to_int_or_none(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    # Handle textual notice periods like "2 months" -> 60
    months_match = re.match(r"(\d+)\s*months?", text, re.IGNORECASE)
    if months_match:
        return int(months_match.group(1)) * 30
    weeks_match = re.match(r"(\d+)\s*weeks?", text, re.IGNORECASE)
    if weeks_match:
        return int(weeks_match.group(1)) * 7
    days_match = re.match(r"(\d+)\s*days?", text, re.IGNORECASE)
    if days_match:
        return int(days_match.group(1))
    cleaned = re.sub(r"[^0-9]", "", text)
    if cleaned:
        try:
            return int(cleaned)
        except ValueError:
            pass
    return None


def _truncate(value: Optional[str], max_len: int = 255) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s[:max_len] if len(s) > max_len else s


# ──────────────────────────────────────────────
# Core: map a row to candidate kwargs
# ──────────────────────────────────────────────

def _has_transportation_to_bool(value: Any) -> Optional[bool]:
    """Application.has_transportation is bool; import mapping still uses TransportationAvailability."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, TransportationAvailability):
        return value is not TransportationAvailability.no
    return None


def _split_profile_and_application(mapped: Dict[str, Any]) -> tuple[Dict[str, Any], Dict[str, Any]]:
    """
    Split flat row kwargs (legacy candidate-shaped) into CandidateProfile and Application dicts.
    Nationality is stored on ``applications.nationality`` (not custom_fields).
    """
    profile: Dict[str, Any] = {
        "organization_id": mapped["organization_id"],
        "import_session_id": mapped["import_session_id"],
    }
    for k in ("full_name", "email", "date_of_birth"):
        if k in mapped:
            profile[k] = mapped[k]

    application: Dict[str, Any] = {"import_session_id": mapped["import_session_id"]}
    for k in APPLICATION_IMPORT_KEYS:
        if k == "import_session_id" or k not in mapped:
            continue
        v = mapped[k]
        if k == "has_transportation":
            v = _has_transportation_to_bool(v)
        application[k] = v

    custom = dict(mapped.get("custom_fields") or {})
    if mapped.get("raw_import_data") is not None:
        custom["_raw_import_data"] = mapped["raw_import_data"]

    reserved_for_split = (
        set(profile.keys())
        | APPLICATION_IMPORT_KEYS
        | {"organization_id", "raw_import_data", "custom_fields"}
    )
    for k, v in mapped.items():
        if k in reserved_for_split:
            continue
        if v is not None:
            custom[k] = _to_json_safe(v)

    application["custom_fields"] = custom
    return profile, application


# ──────────────────────────────────────────────
# Workbook helpers
# ──────────────────────────────────────────────

async def load_workbook_from_file(file: UploadFile):
    if not file.filename.lower().endswith(".xlsx"):
        raise ValueError("Only .xlsx files are supported.")
    try:
        contents = await file.read()
        workbook = load_workbook(filename=BytesIO(contents), data_only=True)
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(f"Failed to read XLSX file: {exc}") from exc
    if not workbook.sheetnames:
        raise ValueError("The provided XLSX file has no sheets.")
    return workbook, contents


def preview_workbook(workbook, filename: str) -> dict:
    sheets_info = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        non_empty_rows = sum(
            1 for row in sheet.iter_rows(min_row=2, values_only=True)
            if any(c is not None and str(c).strip() for c in row)
        )
        sheets_info.append({
            "name": sheet_name,
            "max_row": sheet.max_row,
            "data_rows": non_empty_rows,
        })
    return {
        "file_name": filename,
        "sheets": sheets_info,
        "total_sheets": len(sheets_info),
    }


def _extract_headers_from_row(sheet, row_idx: int) -> List[str]:
    """Extract header strings from a specific row (strip and blank as empty)."""
    headers = []
    for row in sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True):
        for cell in row:
            if cell is not None and str(cell).strip():
                headers.append(str(cell).strip())
            else:
                headers.append("")
    return headers


def _extract_all_headers_from_sheet(sheet) -> List[str]:
    """Collect unique headers from every table (every header row) in the sheet."""
    seen = set()
    ordered = []
    for row_idx in detect_all_header_rows(sheet):
        for h in _extract_headers_from_row(sheet, row_idx):
            if h and h.lower() not in seen:
                seen.add(h.lower())
                ordered.append(h)
    return ordered


def _extract_all_headers(workbook, sheet_names: List[str]) -> List[str]:
    """Collect unique headers across all tables in all selected sheets."""
    seen = set()
    ordered = []
    for name in sheet_names:
        sheet = workbook[name]
        for h in _extract_all_headers_from_sheet(sheet):
            if h and h.lower() not in seen:
                seen.add(h.lower())
                ordered.append(h)
    return ordered


def _collect_column_values(workbook, sheet_names: List[str], header: str) -> List[Any]:
    """Collect sample values from a specific column across all tables in all sheets."""
    values = []
    header_lower = header.strip().lower()
    for name in sheet_names:
        sheet = workbook[name]
        header_row_indices = detect_all_header_rows(sheet)
        max_row = getattr(sheet, "max_row", 0) or 0
        for t_idx, header_row_idx in enumerate(header_row_indices):
            next_header = (
                header_row_indices[t_idx + 1] - 1
                if t_idx + 1 < len(header_row_indices)
                else max_row
            )
            headers_row = list(sheet.iter_rows(
                min_row=header_row_idx, max_row=header_row_idx, values_only=True
            ))[0]
            col_idx = None
            for i, h in enumerate(headers_row):
                if h and str(h).strip().lower() == header_lower:
                    col_idx = i
                    break
            if col_idx is None:
                continue
            for row in sheet.iter_rows(
                min_row=header_row_idx + 1, max_row=next_header, values_only=True
            ):
                if col_idx < len(row) and row[col_idx] is not None:
                    values.append(row[col_idx])
                if len(values) >= 100:
                    return values
    return values


class ImportServiceProtocol(Protocol):
    async def load_workbook_from_file(self, file: UploadFile): ...
    def preview_workbook(self, workbook, filename: str) -> dict: ...
    async def analyze_workbook(
        self,
        workbook,
        contents: bytes,
        filename: str,
        sheet_names: List[str],
        org_id: int,
        user_id: int,
    ) -> dict: ...
    def check_import_name_conflicts(
        self,
        *,
        org_id: int,
        filename: str,
        sheet_names: List[str],
    ) -> dict: ...
    def confirm_and_import(
        self,
        session_id: int,
        confirmed_mappings: Dict[str, str],
        new_custom_fields: List[Dict[str, Any]],
        skip_columns: List[str],
        sheet_names: List[str],
        org_id: int,
    ) -> dict: ...


class ImportService:
    def __init__(
        self,
        import_repo: ImportRepositoryProtocol,
        lookup_service: LookupServiceProtocol,
        custom_field_type_service: CustomFieldTypeServiceProtocol,
        column_normalizer_service: ColumnNormalizerServiceProtocol,
    ) -> None:
        self._import_repo = import_repo
        self._lookup_service = lookup_service
        self._custom_field_type_service = custom_field_type_service
        self._column_normalizer_service = column_normalizer_service

    def _map_row(
        self,
        row_data: Dict[str, Any],
        column_mappings: Dict[str, str],
        org_id: int,
    ) -> Dict[str, Any]:
        candidate_kwargs: Dict[str, Any] = {}
        custom_fields: Dict[str, Any] = {}
        raw_import_data = {k: _to_json_safe(v) for k, v in row_data.items()}

        for excel_header, raw_value in row_data.items():
            header_lower = excel_header.strip().lower()
            db_col = column_mappings.get(header_lower)
            if not db_col:
                continue

            if db_col.startswith("custom:"):
                cf_key = db_col[len("custom:"):]
                custom_fields[cf_key] = _to_json_safe(raw_value)
                continue

            if db_col in LOOKUP_COLUMN_MAP:
                category_code = LOOKUP_COLUMN_MAP[db_col]
                resolved_id = self._lookup_service.resolve_lookup_value(
                    category_code,
                    org_id,
                    str(raw_value) if raw_value else "",
                )
                if resolved_id:
                    candidate_kwargs[db_col] = resolved_id
                continue

            if db_col == "is_open_for_relocation":
                val = _to_relocation_openness_or_none(raw_value)
                if val is not None:
                    candidate_kwargs[db_col] = val
            elif db_col == "has_transportation":
                val = _to_transportation_availability_or_none(raw_value)
                if val is not None:
                    candidate_kwargs[db_col] = val
            elif db_col in BOOLEAN_COLUMNS:
                val = _to_bool_or_none(raw_value)
                if val is not None:
                    candidate_kwargs[db_col] = val
            elif db_col in DECIMAL_COLUMNS:
                val = _to_decimal_or_none(raw_value)
                if val is not None:
                    candidate_kwargs[db_col] = val
            elif db_col in INT_COLUMNS:
                val = _to_int_or_none(raw_value)
                if val is not None:
                    candidate_kwargs[db_col] = val
            elif db_col in DATE_COLUMNS:
                val = _to_date_or_none(raw_value)
                if val is not None:
                    candidate_kwargs[db_col] = val
            elif db_col == "applied_at":
                if isinstance(raw_value, datetime):
                    candidate_kwargs[db_col] = raw_value
                elif isinstance(raw_value, str):
                    try:
                        candidate_kwargs[db_col] = datetime.fromisoformat(raw_value.strip())
                    except ValueError:
                        pass
            elif db_col == "tech_stack":
                if isinstance(raw_value, str):
                    candidate_kwargs[db_col] = [s.strip() for s in raw_value.split(",") if s.strip()]
                elif isinstance(raw_value, list):
                    candidate_kwargs[db_col] = raw_value
            elif db_col == "email":
                email_str = str(raw_value).strip() if raw_value else ""
                if "@" in email_str and "." in email_str:
                    candidate_kwargs[db_col] = _truncate(email_str, 320)
            elif db_col == "nationality":
                if raw_value is not None and str(raw_value).strip():
                    candidate_kwargs[db_col] = _truncate(str(raw_value), 100)
            elif db_col in STRING_COLUMNS:
                if raw_value is not None and str(raw_value).strip():
                    candidate_kwargs[db_col] = _truncate(str(raw_value), 255)

        candidate_kwargs["custom_fields"] = custom_fields if custom_fields else {}
        candidate_kwargs["raw_import_data"] = raw_import_data
        return candidate_kwargs

    def _process_sheet(
        self,
        sheet,
        sheet_name: str,
        col_map: Dict[str, str],
        org_id: int,
        session_id: int,
        skip_headers_lower: Optional[Set[str]] = None,
    ) -> dict:
        skipped = skip_headers_lower or set()
        header_row_indices = detect_all_header_rows(sheet)
        if not header_row_indices:
            header_row_indices = [1]

        created = 0
        skipped_empty = 0
        skipped_duplicates = 0
        row_errors = []
        max_row = getattr(sheet, "max_row", 0) or 0

        for t_idx, header_row_idx in enumerate(header_row_indices):
            next_header = (
                header_row_indices[t_idx + 1]
                if t_idx + 1 < len(header_row_indices)
                else max_row + 1
            )
            data_start = header_row_idx + 1
            data_end = next_header - 1
            if data_end < data_start:
                continue

            headers_row = list(
                sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True)
            )[0]
            headers = [str(h).strip() if h else "" for h in headers_row]

            for row_idx, row in enumerate(
                sheet.iter_rows(min_row=data_start, max_row=data_end, values_only=True),
                start=data_start,
            ):
                row_dict = {}
                for i, h in enumerate(headers):
                    if h and i < len(row):
                        if str(h).strip().lower() in skipped:
                            continue
                        row_dict[h] = row[i]

                if not any(v is not None and str(v).strip() for v in row_dict.values()):
                    skipped_empty += 1
                    continue

                try:
                    mapped = self._map_row(row_dict, col_map, org_id)
                    mapped["organization_id"] = org_id
                    mapped["import_session_id"] = session_id
                    profile_kw, app_kw = _split_profile_and_application(mapped)
                    self._import_repo.insert_imported_candidate_row(profile_kw, app_kw)
                    created += 1
                except Exception as exc:
                    row_errors.append({"row_index": row_idx, "error": str(exc)})

        return {
            "sheet_name": sheet_name,
            "created": created,
            "skipped_empty": skipped_empty,
            "skipped_duplicates": skipped_duplicates,
            "row_errors": row_errors,
        }

    async def load_workbook_from_file(self, file: UploadFile):
        return await load_workbook_from_file(file)

    def preview_workbook(self, workbook, filename: str) -> dict:
        return preview_workbook(workbook, filename)

    async def analyze_workbook(
        self,
        workbook,
        contents: bytes,
        filename: str,
        sheet_names: List[str],
        org_id: int,
        user_id: int,
    ) -> dict:
        session = self._import_repo.create_pending_import_session(
            org_id=org_id,
            user_id=user_id,
            original_filename=filename,
        )
        _save_temp_file(session.id, contents)

        all_headers = _extract_all_headers(workbook, sheet_names)
        norm_result = await self._column_normalizer_service.normalize_columns(
            all_headers,
            org_id,
            use_llm=True,
        )
        self._import_repo.commit_import_transaction()

        def _serialize_mapping(m: ColumnMapping) -> dict:
            return {
                "excel_header": m.excel_header,
                "db_column": m.db_column,
                "db_column_label": COLUMN_LABELS.get(m.db_column, m.db_column) if m.db_column else None,
                "confidence": m.confidence,
                "source": m.source,
            }

        already_mapped = {
            m.db_column
            for m in [*norm_result.matched, *norm_result.suggested]
            if m.db_column
        }
        available = self._column_normalizer_service.get_available_columns(org_id)
        return {
            "session_id": session.id,
            "filename": filename,
            "sheets": sheet_names,
            "matched_columns": [_serialize_mapping(m) for m in norm_result.matched],
            "suggested_columns": [_serialize_mapping(m) for m in norm_result.suggested],
            "unmatched_columns": [_serialize_mapping(m) for m in norm_result.unmatched],
            "available_columns": available,
            "already_mapped": list(already_mapped),
        }

    def check_import_name_conflicts(
        self,
        *,
        org_id: int,
        filename: str,
        sheet_names: List[str],
    ) -> dict:
        normalized_filename = str(filename or "").strip()
        normalized_sheets = [str(s).strip() for s in (sheet_names or []) if str(s).strip()]
        filename_lower = normalized_filename.lower() if normalized_filename else ""
        filename_exists = self._import_repo.import_filename_exists(org_id, filename_lower)

        existing_sheets_map: dict[str, str] = {}
        if filename_lower:
            for sheet_val in self._import_repo.distinct_import_sheet_strings_for_filename(
                org_id, filename_lower
            ):
                for part in sheet_val.split(","):
                    p = part.strip()
                    if p:
                        existing_sheets_map[p.lower()] = p

        duplicate_sheets = []
        for sheet in normalized_sheets:
            match = existing_sheets_map.get(sheet.lower())
            if match:
                duplicate_sheets.append(match)
        duplicate_sheets = list(dict.fromkeys(duplicate_sheets))

        return {
            "filename_exists": filename_exists,
            "duplicate_sheets": duplicate_sheets,
            "has_duplicates": bool(duplicate_sheets),
        }

    def confirm_and_import(
        self,
        session_id: int,
        confirmed_mappings: Dict[str, str],
        new_custom_fields: List[Dict[str, Any]],
        skip_columns: List[str],
        sheet_names: List[str],
        org_id: int,
    ) -> dict:
        session = self._import_repo.get_import_session_by_id(session_id)
        if not session:
            raise ValueError(f"Import session {session_id} not found.")
        if session.status != "pending":
            raise ValueError(f"Session is already {session.status}.")

        self._import_repo.mark_import_session_processing(session)
        workbook = _load_temp_file(session_id)

        for cf_spec in new_custom_fields:
            header = cf_spec["header"]
            label = cf_spec.get("label", header)
            values = _collect_column_values(workbook, sheet_names, header)
            cfd, _ = self._custom_field_type_service.create_custom_field(org_id, label, values)
            confirmed_mappings[header.strip().lower()] = f"custom:{cfd.field_key}"

        col_map = {header.strip().lower(): db_col for header, db_col in confirmed_mappings.items()}
        skip_lower: Set[str] = {
            str(h).strip().lower()
            for h in (skip_columns or [])
            if h is not None and str(h).strip()
        }
        for hl in skip_lower:
            col_map.pop(hl, None)

        total_created = 0
        total_skipped_empty = 0
        total_skipped_duplicates = 0
        total_errors = 0
        sheet_results = []
        all_row_errors = []

        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            result = self._process_sheet(sheet, sheet_name, col_map, org_id, session.id, skip_lower)
            total_created += result["created"]
            total_skipped_empty += result["skipped_empty"]
            total_skipped_duplicates += result["skipped_duplicates"]
            total_errors += len(result["row_errors"])
            sheet_results.append(result)
            all_row_errors.extend({**e, "sheet": sheet_name} for e in result["row_errors"])

        sheet_labels = [str(s).strip() for s in sheet_names if s is not None and str(s).strip()]
        import_sheet_val = ", ".join(dict.fromkeys(sheet_labels))[:255] if sheet_labels else None
        self._import_repo.complete_import_session_and_commit(
            session,
            import_sheet=import_sheet_val,
            total_rows=total_created + total_skipped_empty + total_skipped_duplicates + total_errors,
            imported_rows=total_created,
            skipped_rows=total_skipped_empty + total_skipped_duplicates,
            error_rows=total_errors,
            completed_at=datetime.utcnow(),
            summary={"sheets": sheet_results, "row_errors": all_row_errors[:100]},
        )
        _cleanup_temp_file(session_id)

        return {
            "session_id": session_id,
            "status": session.status,
            "total_created": total_created,
            "total_skipped_empty_rows": total_skipped_empty,
            "total_skipped_duplicates": total_skipped_duplicates,
            "total_errors": total_errors,
            "sheet_results": sheet_results,
            "row_errors": all_row_errors[:50],
        }
